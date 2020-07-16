# -*- coding: utf-8 -*-
# @Time ： 2020/7/15 21:50
# @Auth ： guoguo
# @File ：lesson7.py
# @QQ ：2369256166
# @weixin : 18434755055

'''
接口自动化步骤:
1.exce1测试用例准备ok,代码自动读收测试数据
2、发送接口请求，得到响应信息
3、断言:实际结果vs预期结果--通过/不通过
4、 写入通过/不通过- excel
'''
import requests
import openpyxl

#读取测试用例得函数
def read_book(filename,sheetname):
    wb=openpyxl.load_workbook(filename)     #加载工作簿  写入工作簿名字
    sheet =wb[sheetname]  #获取表单
    max_row=sheet.max_row   #获取最大行数
    # max_column=sheet.max_column  #获取最大列数
    case_list=[]   #创建一个空列表存放测试用例
    for i in range(2,max_row+1):
        dict1=dict(
        case_id = sheet.cell(row=i, column=1).value,
        url=sheet.cell(row=i,column=5) .value,
        data=sheet.cell(row=i,column=6) .value,
        expect=sheet.cell(row=i,column=7) .value)
        case_list.append(dict1)   #每循环一次，把读取到的字典数据放进list中
    return case_list

#执行接口的函数
def guoguo(url,data):
    heads_reg={"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}#请求头--字典格式

    res=requests.post(url=url,json=data,headers=heads_reg,)  #接收post方法的结果
    response=res.json()
    return response

#写入结果
def write_result(filename,sheetname,row,column,final_reult):
    wb=openpyxl.load_workbook(filename)
    sheet =wb[sheetname]
    cell=sheet.cell(row=row,column=column) .value=final_reult #直接写入结果
    wb.save(filename)   #保存



#执行测试用例并回写实际结果
def execute_fun(filename,sheetname):
    cases=read_book(filename,sheetname)    #调用读取测试用例，保存到一个变量中
    for case in cases:
        case_id=case.get('case_id')  # 取id
        url=case.get('url')
        data =eval(case.get('data'))
        expect =eval( case.get('expect'))   #获取预期结果
        expect_msg=expect.get('msg')  #获取我们预期结果中得msg信息
        real_result=guoguo(url=url,data=data)   #调用发送接口请求函数,返回结果用变量real_result接受
        real_msg=real_result.get('msg')   #获取实际结果中的msg信息
        print('预期结果中的msg:{}'.format(expect_msg))
        print('实际结果中的msg:{}'.format(real_msg))
        if real_msg==expect_msg:
            print('第{}条用例通过'.format(case_id))
            final_re='Passed'
        else:
            print('第{}条用例不通过'.format(case_id))
            final_re = 'Failed'
        write_result(filename,sheetname,case_id+1,8,final_re)
        print('*'*25)

execute_fun('test_case_api.xlsx','login')
